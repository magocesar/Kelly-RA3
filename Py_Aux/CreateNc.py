import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
import time

class CreateNc:
    def __init__(self, perguntas, respostas, justificativas, gravidade, nome_proj, responsavel_proj, rqa_proj, acoes_corretivas, superior):
        self.perguntas = perguntas
        self.respostas = respostas
        self.justificativas = justificativas
        self.gravidade = gravidade
        self.nome_proj = nome_proj
        self.responsavel_proj = responsavel_proj
        self.rqa_proj = rqa_proj
        self.len_nc = 0
        self.dir = None
        self.acoes_corretivas = acoes_corretivas
        self.superior = superior

    def start(self):

        if 0 not in self.respostas:

            #Caso exista algum diretório de NC desse projeto, apagar
            if os.path.exists(f"./ExcelNc/{self.nome_proj}"):
                for file in os.listdir(f"./ExcelNc/{self.nome_proj}"):
                    os.remove(f"./ExcelNc/{self.nome_proj}/{file}")
                os.rmdir(f"./ExcelNc/{self.nome_proj}")
            return None, -1

        if self.createFolder() == False:
            return False, False

        self.CreateExcel()

        return self.dir, self.len_nc
        
    def createFolder(self):
        self.dir = f"./ExcelNc/{self.nome_proj}"

        try:
            if os.path.exists(self.dir):
                for file in os.listdir(self.dir):
                    os.remove(f"{self.dir}/{file}")
            else:
                os.makedirs(self.dir)
        except:
            print("Error")
            return False
    
    def CreateExcel(self):
        #Selecionar Perguntas de Foram respondidas com Não
        for i in range(len(self.respostas)):
            if(self.respostas[i] == 0):
                self.NcCreate(self.perguntas[i], self.justificativas[i], self.gravidade[i], self.acoes_corretivas[i])
                self.len_nc += 1
                
    def NcCreate(self, pergunta, justificativa, gravidade, acao):
        wb = Workbook()
        ws = wb.active

        ws.append(["Nome do Projeto", "Responsável do Projeto", "RQA do Projeto", "Pergunta", "Resposta", "Gravidade", "Justificativa", "Ação Corretiva", "Data Limite", "Superior Imediato"])

        #Gravidade:
        #'alta / 3 dias';
        #'média / 2 dias'
        #'baixa / 1 dia'

        if(gravidade == 'Alta'):
            data_limite = time.strftime("%d/%m/%Y", time.localtime(time.time() + 3*24*60*60))
        elif(gravidade == 'Média'):
            data_limite = time.strftime("%d/%m/%Y", time.localtime(time.time() + 2*24*60*60))
        elif(gravidade == 'Baixa'):
            data_limite = time.strftime("%d/%m/%Y", time.localtime(time.time() + 1*24*60*60))


        ws.append([self.nome_proj, self.responsavel_proj, self.rqa_proj, pergunta, 'Não', gravidade, justificativa, acao, data_limite, self.superior])

        tab = Table(displayName="Table1", ref=f"A1:J2")

        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        
        tab.tableStyleInfo = style

        ws.add_table(tab)

        for i in range(1, 3):
            ws[f'A{i}'].alignment = Alignment(wrap_text=True)
            ws[f'B{i}'].alignment = Alignment(wrap_text=True)
            ws[f'C{i}'].alignment = Alignment(wrap_text=True)
            ws[f'D{i}'].alignment = Alignment(wrap_text=True)
            ws[f'E{i}'].alignment = Alignment(wrap_text=True)
            ws[f'F{i}'].alignment = Alignment(wrap_text=True)
            ws[f'G{i}'].alignment = Alignment(wrap_text=True)
            ws[f'H{i}'].alignment = Alignment(wrap_text=True)
            ws[f'I{i}'].alignment = Alignment(wrap_text=True)
            ws[f'J{i}'].alignment = Alignment(wrap_text=True)


        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 40


        wb.save(f"{self.dir}/Nc{self.len_nc + 1}.xlsx")

