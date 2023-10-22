from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from copy import copy
from openpyxl.chart import PieChart, Reference
from time import time
import os

class CheckListCreator:
    def __init__(self, perguntas, respostas, nome_proj):
        self.perguntas = perguntas
        self.respostas = respostas
        self.nome_proj = nome_proj
        self.template = Workbook()

    def start(self):
        self.createFolder()
        self._create_checklist()
        self.create_metrics()
        self.insert_special_table()
        return self.save()
    
    def createFolder(self):
        try:
            if not os.path.exists("./ExcelCheckList"):
                os.makedirs("./ExcelCheckList")
        except:
            print("Error")
            return False

    def _create_checklist(self):
        ws = self.template.active

        tabela = []

        for i in range(len(self.perguntas)):
            tabela.append([self.perguntas[i], list(map(lambda x: 'Sim' if x == 1 else 'Não', self.respostas))[i]])

        ws.append(['Pergunta', 'Resposta'])

        for linha in tabela:
            ws.append(linha)

        tab = Table(displayName="Table1", ref=f"A1:B{len(self.perguntas) + 1}")

        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        
        tab.tableStyleInfo = style

        ws.add_table(tab)

        for i in range(1, len(self.perguntas) + 1):
            ws[f'A{i}'].alignment = Alignment(wrap_text=True)
            ws[f'B{i}'].alignment = Alignment(wrap_text=True)

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
    
    def create_metrics(self):
        ws = self.template.active

        numero_perg = len(self.perguntas)
        numero_nc = len(list(filter(lambda x: x == 0, self.respostas)))
        numero_conforme = len(list(filter(lambda x: x == 1, self.respostas)))
        porcentagem_aderencia = 100 - round((numero_nc / numero_perg) * 100, 2)

        ws["D1"] = "Número de perguntas: "
        ws["E1"] = "Número de conformidades: "
        ws["F1"] = "Número de não conformidades: "
        ws["G1"] = "Porcentagem de aderência: "

        tab = Table(displayName="Table2", ref=f"D1:G2")

        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        
        tab.tableStyleInfo = style

        ws.add_table(tab)

        ws["D2"] = numero_perg
        ws["E2"] = numero_conforme
        ws["F2"] = numero_nc
        ws["G2"] = f"{porcentagem_aderencia}%"

        for col in ws:
            for cell in col:
                alignment_obj = copy(cell.alignment)
                alignment_obj.horizontal = 'center'
                cell.alignment = alignment_obj

        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30

    def insert_special_table(self):
        ws = self.template.active

        pie = PieChart()

        labels = Reference(ws, min_col=5, min_row=1, max_row=2, max_col=6)
        
        data = Reference(ws, min_col=5, min_row=2, max_col=6, max_row=2)

        pie.add_data(data, titles_from_data=True)

        pie.set_categories(labels)

        pie.title = "Gráfico de Conformidades: " + f"{self.nome_proj}"

        ws.add_chart(pie, "D4")

    def save(self):
        currentTime = time()
        self.template.save(f"ExcelCheckList/CheckList_{self.nome_proj}_{currentTime}.xlsx")
        return f"Excel/CheckList_{self.nome_proj}_{currentTime}.xlsx"